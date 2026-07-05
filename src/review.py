"""
Interactive/manual review tools for the discovery backlog (a_examiner).

Provides preview (multi-format), column tagging, RM counting, and the CLI
revue_manuelle_a_examiner() flow. Used by dashboard.py (web) and cli.py (terminal).

No module-level dependency on discover.py — circular-import-safe thanks to lazy
imports inside revue_manuelle_a_examiner().
"""

import bz2
import csv
import gzip
import io
import json
import os
import sys
import warnings
import zipfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from conf.communes_rm import CODES_POSTAUX_RM
from connectors.analyseurs import _detecter_delimiteur, _format_analysable
from connectors.download import _telecharger
from connectors.sirene import obtenir_sirens_rm
from filters.discovery import trouver_ressource_analysable
from filters.geographic import (
    normaliser,
    est_circonscription_rm,
    est_valeur_commune_rm,
    est_epci_rm,
    est_point_rm,
    est_adresse_rm,
)

_TYPES_VARIABLES = [
    ("commune", "Commune (code INSEE/IRIS, code postal, ou nom)"),
    ("cp", "Code postal seul"),
    ("epci", "EPCI (code SIREN de l'intercommunalité)"),
    ("latlon", "Latitude / longitude (WKT \"POINT/POLYGON(...)\" ou géométrie GeoJSON \"geom\")"),
    ("siren", "SIREN / SIRET"),
    ("adresse", "Adresse complète"),
    ("circonscription", "Circonscription législative (Assemblée nationale)"),
]


def _decoder_apercu_csv(contenu: bytes):
    if contenu[:5] in (b"%PDF-", b"PK\x03\x04", b"\x1f\x8b\x08"):
        return None
    debut = contenu[:100].lstrip().lower()
    if debut.startswith((b"<!doctype", b"<html")):
        return None

    texte = contenu.decode("utf-8-sig", errors="replace")
    if texte.count("�") > 10:
        texte = contenu.decode("latin-1")
    sample = texte[:4096]
    delimiteur = _detecter_delimiteur(sample)

    premiere_ligne = texte.split("\n")[0]
    premiere_norm = normaliser(premiere_ligne.split(",")[0].split(";")[0])
    if premiere_norm in ("colonne", "column", "champ", "field", "variable"):
        return "__DICTIONNAIRE__"

    reader = csv.DictReader(io.StringIO(texte, newline=""), delimiter=delimiteur)
    entetes = list(reader.fieldnames or [])
    return texte, delimiteur, entetes


_PREVIEW_BORNE = 1024 * 1024
_SNIFF_BORNE = 65536


def _obtenir_bytes_ressource(ressource: dict, fmt: str) -> tuple:
    url = ressource.get("url", "")

    if fmt == "csv":
        chemin, _, _, erreur = _telecharger(url, verbose=False)
        if erreur:
            return False, f"Échec du téléchargement : {erreur}", False
        with open(chemin, "rb") as f:
            return True, "csv", f.read()

    if fmt == "xlsx":
        chemin, _, _, erreur = _telecharger(url, verbose=False)
        if erreur:
            return False, f"Échec du téléchargement : {erreur}", False
        with open(chemin, "rb") as f:
            return True, "xlsx", f.read()

    if fmt == "geojson":
        chemin, _, _, erreur = _telecharger(url, verbose=False)
        if erreur:
            return False, f"Échec du téléchargement : {erreur}", False
        with open(chemin, "rb") as f:
            return True, "geojson", f.read()

    return False, "Format non pris en charge pour la revue manuelle.", True


def _finaliser_apercu_csv_bytes(contenu: bytes) -> tuple:
    decode = _decoder_apercu_csv(contenu)
    if decode is None:
        return False, "Fichier binaire ou réponse HTML — impossible d'afficher un aperçu.", False
    if decode == "__DICTIONNAIRE__":
        return False, "Ressource détectée comme dictionnaire de colonnes, pas des données.", True
    texte, delimiteur, entetes = decode
    if not entetes:
        return False, "Aucune colonne détectée.", True
    lignes_apercu = list(csv.DictReader(io.StringIO(texte, newline=""), delimiter=delimiteur))[:5]
    return True, entetes, lignes_apercu


def _sniffer_delimiteur(preambule: bytes) -> tuple:
    decode = _decoder_apercu_csv(preambule)
    if decode is None or decode == "__DICTIONNAIRE__" or not decode[2]:
        return False, "Contenu illisible pour le filtrage."
    _, delimiteur, _ = decode
    return True, delimiteur


def _apercu_csv_gz(chemin: str) -> tuple:
    try:
        with gzip.open(chemin, "rb") as f:
            contenu = f.read(_PREVIEW_BORNE)
    except Exception as e:
        return False, f"Erreur de décompression GZ : {e}", False
    return _finaliser_apercu_csv_bytes(contenu)


def _lignes_csv_gz(chemin: str) -> tuple:
    try:
        with gzip.open(chemin, "rb") as f:
            preambule = f.read(_SNIFF_BORNE)
    except Exception as e:
        return False, f"Erreur de décompression GZ : {e}"
    ok, delimiteur_ou_message = _sniffer_delimiteur(preambule)
    if not ok:
        return False, delimiteur_ou_message
    delimiteur = delimiteur_ou_message

    def _rows():
        with gzip.open(chemin, "rt", encoding="utf-8-sig", errors="replace", newline="") as f:
            yield from csv.DictReader(f, delimiter=delimiteur)

    return True, _rows()


def _apercu_csv_bz2(chemin: str) -> tuple:
    try:
        with bz2.open(chemin, "rb") as f:
            contenu = f.read(_PREVIEW_BORNE)
    except Exception as e:
        return False, f"Erreur de décompression BZ2 : {e}", False
    return _finaliser_apercu_csv_bytes(contenu)


def _lignes_csv_bz2(chemin: str) -> tuple:
    try:
        with bz2.open(chemin, "rb") as f:
            preambule = f.read(_SNIFF_BORNE)
    except Exception as e:
        return False, f"Erreur de décompression BZ2 : {e}"
    ok, delimiteur_ou_message = _sniffer_delimiteur(preambule)
    if not ok:
        return False, delimiteur_ou_message
    delimiteur = delimiteur_ou_message

    def _rows():
        with bz2.open(chemin, "rt", encoding="utf-8-sig", errors="replace", newline="") as f:
            yield from csv.DictReader(f, delimiter=delimiteur)

    return True, _rows()


def _membre_csv_zip(chemin: str) -> str | None:
    try:
        with zipfile.ZipFile(chemin) as zf:
            for n in zf.namelist():
                if not n.startswith("__MACOSX") and n.lower().endswith(".csv"):
                    return n
    except zipfile.BadZipFile:
        return None
    return None


def _apercu_csv_zip_membre(chemin: str, membre: str) -> tuple:
    try:
        with zipfile.ZipFile(chemin) as zf, zf.open(membre) as f:
            contenu = f.read(_PREVIEW_BORNE)
    except Exception as e:
        return False, f"Erreur de lecture ZIP : {e}", False
    return _finaliser_apercu_csv_bytes(contenu)


def _lignes_csv_zip_membre(chemin: str, membre: str) -> tuple:
    try:
        with zipfile.ZipFile(chemin) as zf, zf.open(membre) as f:
            preambule = f.read(_SNIFF_BORNE)
    except Exception as e:
        return False, f"Erreur de lecture ZIP : {e}"
    ok, delimiteur_ou_message = _sniffer_delimiteur(preambule)
    if not ok:
        return False, delimiteur_ou_message
    delimiteur = delimiteur_ou_message

    def _rows():
        with zipfile.ZipFile(chemin) as zf, zf.open(membre) as f_bin:
            f_texte = io.TextIOWrapper(f_bin, encoding="utf-8-sig", errors="replace", newline="")
            yield from csv.DictReader(f_texte, delimiter=delimiteur)

    return True, _rows()


def _entetes_et_apercu_xlsx(contenu: bytes) -> tuple:
    try:
        import openpyxl
    except ImportError:
        return False, "openpyxl non installé — pip install openpyxl.", False
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(io.BytesIO(contenu), read_only=True, data_only=True)
        ws = wb.active
        lignes = list(ws.iter_rows(values_only=True, max_row=6))
        wb.close()
    except Exception as e:
        return False, f"Erreur de lecture XLSX : {e}", False
    if not lignes:
        return False, "Fichier XLSX vide.", True
    entetes = [str(c or "").strip() for c in lignes[0]]
    if not any(entetes):
        return False, "Aucune colonne détectée dans le XLSX.", True
    lignes_apercu = [dict(zip(entetes, (str(v or "").strip() for v in row))) for row in lignes[1:6]]
    return True, entetes, lignes_apercu


def _toutes_lignes_xlsx(contenu: bytes) -> tuple:
    try:
        import openpyxl
    except ImportError:
        return False, "openpyxl non installé — pip install openpyxl."
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(io.BytesIO(contenu), read_only=True, data_only=True)
        ws = wb.active
        lignes = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return False, f"Erreur de lecture XLSX : {e}"
    if not lignes:
        return False, "Fichier XLSX vide."
    entetes = [str(c or "").strip() for c in lignes[0]]
    rows = (dict(zip(entetes, (str(v or "").strip() for v in row))) for row in lignes[1:])
    return True, rows


def _entetes_et_apercu_geojson(contenu: bytes) -> tuple:
    try:
        data = json.loads(contenu.decode("utf-8", errors="replace"))
    except Exception as e:
        return False, f"Erreur de parsing JSON : {e}", False
    features = data.get("features", [])
    if not features:
        return False, "Aucune feature GeoJSON trouvée.", True
    entetes = list((features[0].get("properties") or {}).keys())
    if not entetes:
        return False, "Aucune propriété exploitable (géométrie seule).", True
    lignes_apercu = [(f.get("properties") or {}) for f in features[:5]]
    return True, entetes, lignes_apercu


def _toutes_lignes_geojson(contenu: bytes) -> tuple:
    try:
        data = json.loads(contenu.decode("utf-8", errors="replace"))
    except Exception as e:
        return False, f"Erreur de parsing JSON : {e}"
    features = data.get("features", [])
    return True, (f.get("properties") or {} for f in features)


def analyser_apercu_revue(ressource: dict) -> tuple:
    fmt = _format_analysable(ressource)
    if fmt is None:
        return False, "Aucune ressource dans un format pris en charge pour la revue manuelle.", True
    url = ressource.get("url", "")

    if fmt == "parquet":
        try:
            import pyarrow.parquet as pq
            import fsspec
        except ImportError:
            return False, "pyarrow/fsspec non installés — pip install pyarrow fsspec.", False
        try:
            fs, fpath = fsspec.url_to_fs(url)
            with fs.open(fpath, "rb") as f:
                pf = pq.ParquetFile(f)
                entetes = [field.name for field in pf.schema_arrow]
                try:
                    lignes_apercu = next(pf.iter_batches(batch_size=5)).to_pylist()
                except StopIteration:
                    lignes_apercu = []
        except Exception as e:
            return False, f"Erreur de lecture Parquet : {e}", False
        if not entetes:
            return False, "Aucune colonne détectée dans le Parquet.", True
        return True, fmt, entetes, lignes_apercu

    if fmt == "gz":
        chemin, _, _, erreur = _telecharger(url, verbose=False)
        if erreur:
            return False, f"Échec du téléchargement : {erreur}", False
        resultat = _apercu_csv_gz(chemin)
        if not resultat[0]:
            return resultat
        _, entetes, lignes_apercu = resultat
        return True, fmt, entetes, lignes_apercu

    if fmt == "bz2":
        chemin, _, _, erreur = _telecharger(url, verbose=False)
        if erreur:
            return False, f"Échec du téléchargement : {erreur}", False
        resultat = _apercu_csv_bz2(chemin)
        if not resultat[0]:
            return resultat
        _, entetes, lignes_apercu = resultat
        return True, fmt, entetes, lignes_apercu

    if fmt == "zip":
        chemin, _, _, erreur = _telecharger(url, verbose=False, plafond_mo=None)
        if erreur:
            return False, f"Échec du téléchargement : {erreur}", False
        membre_csv = _membre_csv_zip(chemin)
        if membre_csv:
            resultat = _apercu_csv_zip_membre(chemin, membre_csv)
            if not resultat[0]:
                return resultat
            _, entetes, lignes_apercu = resultat
            return True, fmt, entetes, lignes_apercu
        try:
            with zipfile.ZipFile(chemin) as zf:
                membres_geo = [n for n in zf.namelist()
                               if not n.startswith("__MACOSX") and n.lower().endswith(".geojson")]
                if not membres_geo:
                    return False, "Archive ZIP sans fichier CSV ou GeoJSON exploitable.", True
                with zf.open(membres_geo[0]) as f:
                    contenu = f.read()
        except zipfile.BadZipFile:
            return False, "Archive ZIP invalide.", False
        resultat = _entetes_et_apercu_geojson(contenu)
        if not resultat[0]:
            return resultat
        _, entetes, lignes_apercu = resultat
        return True, fmt, entetes, lignes_apercu

    resultat = _obtenir_bytes_ressource(ressource, fmt)
    if not resultat[0]:
        return resultat
    _, sous_format, contenu = resultat

    if sous_format == "csv":
        resultat = _finaliser_apercu_csv_bytes(contenu)
    elif sous_format == "geojson":
        resultat = _entetes_et_apercu_geojson(contenu)
    elif sous_format == "xlsx":
        resultat = _entetes_et_apercu_xlsx(contenu)
    else:
        return False, "Format non pris en charge pour la revue manuelle.", True

    if not resultat[0]:
        return resultat
    _, entetes, lignes_apercu = resultat
    return True, fmt, entetes, lignes_apercu


def analyser_lignes_revue(ressource: dict, fmt: str, colonnes_utiles: list[str] | None = None) -> tuple:
    url = ressource.get("url", "")

    if fmt == "parquet":
        try:
            import pyarrow.parquet as pq
            import fsspec
        except ImportError:
            return False, "pyarrow/fsspec non installés — pip install pyarrow fsspec."
        try:
            fs, fpath = fsspec.url_to_fs(url)
            with fs.open(fpath, "rb") as f:
                pq.ParquetFile(f).schema_arrow
        except Exception as e:
            return False, f"Erreur de lecture Parquet : {e}"

        def _rows():
            with fs.open(fpath, "rb") as f:
                pf = pq.ParquetFile(f)
                for batch in pf.iter_batches(columns=colonnes_utiles or None):
                    yield from batch.to_pylist()

        return True, _rows()

    if fmt == "gz":
        chemin, _, _, erreur = _telecharger(url, verbose=False)
        if erreur:
            return False, f"Échec du téléchargement : {erreur}"
        return _lignes_csv_gz(chemin)

    if fmt == "bz2":
        chemin, _, _, erreur = _telecharger(url, verbose=False)
        if erreur:
            return False, f"Échec du téléchargement : {erreur}"
        return _lignes_csv_bz2(chemin)

    if fmt == "zip":
        chemin, _, _, erreur = _telecharger(url, verbose=False, plafond_mo=None)
        if erreur:
            return False, f"Échec du téléchargement : {erreur}"
        membre_csv = _membre_csv_zip(chemin)
        if membre_csv:
            return _lignes_csv_zip_membre(chemin, membre_csv)
        try:
            with zipfile.ZipFile(chemin) as zf:
                membres_geo = [n for n in zf.namelist()
                               if not n.startswith("__MACOSX") and n.lower().endswith(".geojson")]
                if not membres_geo:
                    return False, "Archive ZIP sans fichier CSV ou GeoJSON exploitable."
                with zf.open(membres_geo[0]) as f:
                    contenu = f.read()
        except zipfile.BadZipFile:
            return False, "Archive ZIP invalide."
        return _toutes_lignes_geojson(contenu)

    resultat = _obtenir_bytes_ressource(ressource, fmt)
    if not resultat[0]:
        return False, resultat[1]
    _, sous_format, contenu = resultat

    if sous_format == "csv":
        decode = _decoder_apercu_csv(contenu)
        if decode is None or decode == "__DICTIONNAIRE__" or not decode[2]:
            return False, "Contenu illisible pour le filtrage."
        texte, delimiteur, _ = decode
        return True, csv.DictReader(io.StringIO(texte, newline=""), delimiter=delimiteur)

    if sous_format == "geojson":
        return _toutes_lignes_geojson(contenu)

    if sous_format == "xlsx":
        return _toutes_lignes_xlsx(contenu)

    return False, "Format non pris en charge pour la revue manuelle."


def _apercu_colonnes(entetes: list[str], lignes: list[dict]) -> None:
    for i, entete in enumerate(entetes, 1):
        valeurs = []
        for ligne in lignes:
            v = str(ligne.get(entete, ""))
            valeurs.append(v[:20] + ("…" if len(v) > 20 else ""))
        print(f"  {i:2d}. {entete[:30]:30s} : {', '.join(valeurs)}")


def _choisir_colonne(entetes: list[str], question: str) -> str:
    while True:
        saisie = input(question).strip()
        if saisie.isdigit() and 1 <= int(saisie) <= len(entetes):
            return entetes[int(saisie) - 1]
        print("    (numéro invalide)")


def _choisir_variable_et_type(entetes: list[str]) -> tuple[str, str, str | None]:
    col1 = _choisir_colonne(
        entetes, "\n  Quelle colonne sert à filtrer Rennes Métropole (numéro) ? ")

    print("\n  Quel type de variable est-ce ?")
    for i, (_, label) in enumerate(_TYPES_VARIABLES, 1):
        print(f"    {i}. {label}")
    while True:
        saisie = input("  Numéro : ").strip()
        if saisie.isdigit() and 1 <= int(saisie) <= len(_TYPES_VARIABLES):
            type_variable = _TYPES_VARIABLES[int(saisie) - 1][0]
            break
        print("    (numéro invalide)")

    col2 = None
    if type_variable == "latlon":
        reponse = input("  La colonne choisie contient-elle déjà lat ET lon combinés "
                         "(ex: '48.11,-1.68', WKT \"POINT(...)\"/\"POLYGON(...)\", ou une "
                         "géométrie GeoJSON \"{coordinates: ...}\") ? (o/n) ").strip().lower()
        if reponse != "o":
            col2 = _choisir_colonne(
                entetes, "  Quelle colonne contient l'autre coordonnée (numéro) ? ")

    return type_variable, col1, col2


def _compter_lignes_variable(rows, type_variable: str, col1: str, col2: str | None = None) -> tuple:
    sirens_rm = obtenir_sirens_rm() if type_variable == "siren" else None
    nb_total, nb_rm = 0, 0
    exemples, premieres_lignes = [], []
    for row in rows:
        try:
            nb_total += 1
            if len(premieres_lignes) < 5:
                premieres_lignes.append(dict(row))
            if type_variable == "commune":
                in_rm = est_valeur_commune_rm(row.get(col1, ""))
            elif type_variable == "cp":
                in_rm = str(row.get(col1, "")).strip() in CODES_POSTAUX_RM
            elif type_variable == "epci":
                in_rm = est_epci_rm(row.get(col1, ""))
            elif type_variable == "latlon":
                lon_val = str(row.get(col2, "")).strip() if col2 else None
                in_rm = est_point_rm(str(row.get(col1, "")).strip(), lon_val)
            elif type_variable == "siren":
                val = str(row.get(col1, "")).strip().replace(" ", "")
                in_rm = val.isdigit() and len(val) in (9, 14) and val[:9] in sirens_rm
            elif type_variable == "adresse":
                in_rm = est_adresse_rm(str(row.get(col1, "")))
            elif type_variable == "circonscription":
                in_rm = est_circonscription_rm(row.get(col1, ""))
            else:
                in_rm = False
            if in_rm:
                nb_rm += 1
                if len(exemples) < 3:
                    exemples.append(dict(row))
        except csv.Error:
            break
    return nb_total, nb_rm, exemples, premieres_lignes


def _construire_champs_manuels(type_variable: str, col1: str, col2: str | None, nb_rm: int) -> dict:
    champs = {"champ_cp": None, "champ_ville": None, "champ_iris": None,
              "champ_adresse": None, "champ_siren": None,
              "champ_epci": None, "champ_lat": None, "champ_lon": None,
              "champ_circonscription": None,
              "nb_rm": nb_rm}
    if type_variable == "commune":
        champs["champ_iris"] = col1
    elif type_variable == "cp":
        champs["champ_cp"] = col1
    elif type_variable == "epci":
        champs["champ_epci"] = col1
    elif type_variable == "latlon":
        champs["champ_lat"] = col1
        champs["champ_lon"] = col2
    elif type_variable == "siren":
        champs["champ_siren"] = col1
    elif type_variable == "adresse":
        champs["champ_adresse"] = col1
    elif type_variable == "circonscription":
        champs["champ_circonscription"] = col1
    return champs


def revue_manuelle_a_examiner() -> None:
    from connectors.datagouv import get_dataset_metadata
    from discover import SEP, _resumer_ligne, charger_decouverte, resoudre_a_examiner

    decouverte = charger_decouverte()
    ids = [e["dataset_id"] for e in decouverte.get("a_examiner", [])
           if e.get("type", "tabulaire") == "tabulaire"]
    if not ids:
        print("Aucun JDD tabulaire en attente d'examen manuel (a_examiner).")
        return

    print(f"{len(ids)} JDD tabulaire(s) en attente d'examen manuel.\n")
    n_ajoutes, n_exclus, n_passes = 0, 0, 0

    for i, did in enumerate(ids, 1):
        entree = next((e for e in decouverte["a_examiner"] if e["dataset_id"] == did), None)
        if entree is None:
            continue

        print(f"\n{SEP}")
        print(f"[{i}/{len(ids)}] {entree['titre'][:70]}")
        print(f"  Organisation : {entree.get('organisation', '')}")
        print(f"  Raison       : {entree.get('raison', '')}")
        print(f"  URL          : {entree.get('url', '')}")

        choix1 = input("\n  (p) passer une fois  (x) exclure définitivement  "
                       "(a) analyser manuellement  (q) quitter la revue ? ").strip().lower()
        if choix1 == "x":
            resoudre_a_examiner(decouverte, did, "exclure")
            print("  → Marqué comme faux positif (exclu définitivement).")
            n_exclus += 1
            continue
        elif choix1 == "q":
            print(f"\nRevue interrompue. {n_ajoutes} ajouté(s), {n_exclus} exclu(s), {n_passes} passé(s).")
            return
        elif choix1 != "a":
            n_passes += 1
            continue

        try:
            metadata = get_dataset_metadata(did)
        except Exception as e:
            print(f"  (impossible de récupérer les métadonnées : {e})")
            n_passes += 1
            continue

        ressource = trouver_ressource_analysable(metadata)
        if ressource is None or _format_analysable(ressource) != "csv":
            fmt_desc = ressource.get("format", "?") if ressource else "aucune ressource"
            print(f"  (type de ressource non supporté pour la revue manuelle : {fmt_desc}"
                  f" — laissé dans le backlog)")
            n_passes += 1
            continue

        chemin, taille_mo, depuis_cache, erreur = _telecharger(ressource["url"], verbose=True)
        if erreur:
            print(f"  (échec du téléchargement : {erreur})")
            n_passes += 1
            continue
        with open(chemin, "rb") as f:
            contenu = f.read()

        decode = _decoder_apercu_csv(contenu)
        if decode is None:
            print("  (fichier binaire ou réponse HTML — impossible d'afficher un aperçu)")
            n_passes += 1
            continue
        if decode == "__DICTIONNAIRE__":
            print("  (ressource détectée comme dictionnaire de colonnes, pas des données — ignorée)")
            n_passes += 1
            continue
        texte, delimiteur, entetes = decode
        if not entetes:
            print("  (aucune colonne détectée)")
            n_passes += 1
            continue

        while True:
            lignes_exemple = list(csv.DictReader(io.StringIO(texte, newline=""), delimiter=delimiteur))[:5]
            print(f"\n  Colonnes disponibles ({len(entetes)}) — valeurs d'exemple sur les 5 premières lignes :")
            _apercu_colonnes(entetes, lignes_exemple)

            type_variable, col1, col2 = _choisir_variable_et_type(entetes)

            rows = csv.DictReader(io.StringIO(texte, newline=""), delimiter=delimiteur)
            nb_total, nb_rm, exemples, _ = _compter_lignes_variable(rows, type_variable, col1, col2)
            print(f"\n  → {nb_rm} ligne(s) RM sur {nb_total}.")
            if exemples:
                print("  Exemples :")
                for ex in exemples:
                    print("    " + _resumer_ligne(ex))

            choix2 = input("\n  (o) ajouter au catalogue  (r) réessayer  "
                           "(p) passer  (x) exclure  (q) quitter la revue ? ").strip().lower()

            if choix2 == "o":
                if nb_rm == 0 and input("  0 ligne RM — ajouter quand même au catalogue ? "
                                        "(o/n) ").strip().lower() != "o":
                    continue
                champs_manuels = _construire_champs_manuels(type_variable, col1, col2, nb_rm)
                resoudre_a_examiner(decouverte, did, "candidat", champs_manuels=champs_manuels)
                print("  → Ajouté au catalogue.")
                n_ajoutes += 1
                break
            elif choix2 == "r":
                continue
            elif choix2 == "x":
                resoudre_a_examiner(decouverte, did, "exclure")
                print("  → Marqué comme faux positif (exclu définitivement).")
                n_exclus += 1
                break
            elif choix2 == "q":
                print(f"\nRevue interrompue. {n_ajoutes} ajouté(s), {n_exclus} exclu(s), {n_passes} passé(s).")
                return
            else:
                n_passes += 1
                break

    print(f"\nFin de la revue manuelle. {n_ajoutes} ajouté(s), {n_exclus} exclu(s), {n_passes} passé(s).")
